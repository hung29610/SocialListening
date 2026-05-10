#!/usr/bin/env python3
"""
Quick Excel Reader for MHC file
"""

import openpyxl
import json
from typing import Dict, List, Any

def read_excel_file(file_path: str) -> Dict[str, Any]:
    """Read and parse Excel file"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        result = {
            'sheet_names': workbook.sheetnames,
            'sheets_data': {},
            'summary': {},
            'json_structure_suggestion': {}
        }
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Find header row (look for row with most text content)
            header_row = None
            headers = []
            
            for row_idx in range(1, 15):  # Check first 15 rows
                row_data = [sheet.cell(row_idx, col).value for col in range(1, 20)]
                text_count = sum(1 for cell in row_data if cell and isinstance(cell, str) and len(str(cell)) > 2)
                
                if text_count >= 5:  # Row with many text cells is likely header
                    header_row = row_idx
                    headers = [cell for cell in row_data if cell is not None][:10]  # Take first 10 non-null headers
                    break
            
            # Get first 10 data rows
            first_10_rows = []
            start_row = (header_row or 1) + 1
            
            for row_idx in range(start_row, min(start_row + 10, sheet.max_row + 1)):
                row_data = []
                for col_idx in range(1, len(headers) + 1):
                    cell_value = sheet.cell(row_idx, col_idx).value
                    row_data.append(cell_value)
                first_10_rows.append(row_data)
            
            # Sheet summary
            non_empty_rows = 0
            for row_idx in range(start_row, sheet.max_row + 1):
                row_has_data = any(sheet.cell(row_idx, col).value for col in range(1, len(headers) + 1))
                if row_has_data:
                    non_empty_rows += 1
            
            result['sheets_data'][sheet_name] = {
                'header_row': header_row,
                'headers': headers,
                'first_10_rows': first_10_rows,
                'dimensions': {
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column
                }
            }
            
            result['summary'][sheet_name] = {
                'total_rows': sheet.max_row,
                'total_columns': sheet.max_column,
                'header_row': header_row,
                'data_rows': non_empty_rows,
                'column_count': len(headers)
            }
        
        # JSON structure suggestion
        result['json_structure_suggestion'] = create_json_structure_suggestion(result)
        
        workbook.close()
        return result
        
    except Exception as e:
        return {'error': f"Failed to read Excel file: {str(e)}"}

def create_json_structure_suggestion(data: Dict[str, Any]) -> Dict[str, Any]:
    """Create JSON structure suggestion based on the data"""
    suggestion = {
        'recommended_structure': {
            'metadata': {
                'filename': 'string',
                'total_sheets': 'number',
                'sheet_names': ['array of strings']
            },
            'services': {
                'sheet_name': {
                    'categories': [
                        {
                            'category_name': 'string',
                            'services': [
                                {
                                    'id': 'number',
                                    'name': 'string',
                                    'details': 'string',
                                    'pricing': {
                                        'type': 'fixed|range|negotiable',
                                        'min_price': 'number|null',
                                        'max_price': 'number|null',
                                        'currency': 'VND',
                                        'raw_text': 'string'
                                    },
                                    'unit': 'string',
                                    'quantity': 'number',
                                    'duration': 'string',
                                    'notes': 'string'
                                }
                            ]
                        }
                    ]
                }
            }
        },
        'field_mapping': {}
    }
    
    # Create field mapping based on detected headers
    for sheet_name, sheet_data in data.get('sheets_data', {}).items():
        headers = sheet_data.get('headers', [])
        mapping = {}
        
        for header in headers:
            if not header:
                continue
            header_lower = str(header).lower()
            
            if 'stt' in header_lower or header_lower.isdigit():
                mapping[header] = 'service_id'
            elif 'dịch vụ' in header_lower or 'service' in header_lower:
                mapping[header] = 'service_name'
            elif 'chi tiết' in header_lower or 'detail' in header_lower:
                mapping[header] = 'service_details'
            elif 'giá' in header_lower or 'price' in header_lower:
                mapping[header] = 'pricing'
            elif 'đơn vị' in header_lower or 'unit' in header_lower:
                mapping[header] = 'unit'
            elif 'số lượng' in header_lower or 'quantity' in header_lower:
                mapping[header] = 'quantity'
            elif 'thời gian' in header_lower or 'time' in header_lower:
                mapping[header] = 'duration'
            elif 'ghi chú' in header_lower or 'note' in header_lower:
                mapping[header] = 'notes'
            else:
                mapping[header] = f"field_{header.replace(' ', '_').replace('\n', '_')}"
        
        suggestion['field_mapping'][sheet_name] = mapping
    
    return suggestion

def main():
    """Main function"""
    file_path = "data/mhc_bao_gia_khung_xlkh.xlsx"
    
    print("Reading Excel file...")
    data = read_excel_file(file_path)
    
    if 'error' in data:
        print(f"Error: {data['error']}")
        return
    
    # Print results
    print("\n" + "="*60)
    print("EXCEL FILE ANALYSIS RESULTS")
    print("="*60)
    
    # 1. Sheet names
    print(f"\n1. SHEET NAMES ({len(data['sheet_names'])} sheets):")
    for i, name in enumerate(data['sheet_names'], 1):
        print(f"   {i}. {name}")
    
    # 2. Column headers for each sheet
    print(f"\n2. COLUMN HEADERS:")
    for sheet_name, sheet_data in data['sheets_data'].items():
        print(f"\n   Sheet: {sheet_name}")
        print(f"   Header Row: {sheet_data['header_row']}")
        print(f"   Headers ({len(sheet_data['headers'])}):")
        for i, header in enumerate(sheet_data['headers'], 1):
            print(f"      {i}. {header}")
    
    # 3. First 10 rows of each sheet
    print(f"\n3. FIRST 10 ROWS OF EACH SHEET:")
    for sheet_name, sheet_data in data['sheets_data'].items():
        print(f"\n   Sheet: {sheet_name}")
        print(f"   Headers: {sheet_data['headers']}")
        print(f"   Data rows:")
        
        for i, row in enumerate(sheet_data['first_10_rows'], 1):
            print(f"      Row {i}: {row}")
            if i >= 5:  # Limit display to first 5 rows for readability
                remaining = len(sheet_data['first_10_rows']) - 5
                if remaining > 0:
                    print(f"      ... and {remaining} more rows")
                break
    
    # 4. Data summary
    print(f"\n4. DATA SUMMARY:")
    for sheet_name, summary in data['summary'].items():
        print(f"\n   Sheet: {sheet_name}")
        print(f"   - Total rows: {summary['total_rows']}")
        print(f"   - Total columns: {summary['total_columns']}")
        print(f"   - Header row: {summary['header_row']}")
        print(f"   - Data rows: {summary['data_rows']}")
        print(f"   - Parsed columns: {summary['column_count']}")
    
    # 5. JSON structure suggestion
    print(f"\n5. JSON STRUCTURE SUGGESTION:")
    print(json.dumps(data['json_structure_suggestion'], indent=2, ensure_ascii=False))
    
    # Save to file
    with open('data/quick_excel_analysis.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    print(f"\n" + "="*60)
    print("Full analysis saved to: data/quick_excel_analysis.json")
    print("="*60)

if __name__ == "__main__":
    main()