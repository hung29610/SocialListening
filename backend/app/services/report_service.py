"""
Report Generation Service - Generate PDF and Excel reports
"""
from datetime import datetime, timedelta
from typing import Dict, Any, List
from io import BytesIO
import os

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference

from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.mention import Mention, AIAnalysis, SentimentScore
from app.models.alert import Alert, AlertSeverity
from app.models.source import Source
from app.models.report import Report, ReportType, ReportStatus


class ReportService:
    """Service for generating reports"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup custom styles for PDF"""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        
        # Heading style
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#374151'),
            spaceAfter=12,
            spaceBefore=12
        ))
    
    async def generate_report(
        self,
        db: AsyncSession,
        report: Report
    ) -> Dict[str, Any]:
        """Generate report based on type"""
        try:
            # Gather data
            data = await self._gather_report_data(db, report)
            
            # Generate PDF
            pdf_path = None
            if report.format in ["pdf", "both"]:
                pdf_path = await self._generate_pdf(report, data)
            
            # Generate Excel
            excel_path = None
            if report.format in ["excel", "both"]:
                excel_path = await self._generate_excel(report, data)
            
            return {
                "success": True,
                "pdf_path": pdf_path,
                "excel_path": excel_path,
                "data": data
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    async def _gather_report_data(
        self,
        db: AsyncSession,
        report: Report
    ) -> Dict[str, Any]:
        """Gather data for report"""
        start_date = report.start_date
        end_date = report.end_date
        
        # Base query for mentions in date range
        mention_query = select(Mention).where(
            and_(
                Mention.collected_at >= start_date,
                Mention.collected_at <= end_date
            )
        )
        
        # Get all mentions
        result = await db.execute(mention_query)
        mentions = result.scalars().all()
        
        # Get AI analysis for mentions
        analysis_query = select(AIAnalysis).where(
            AIAnalysis.mention_id.in_([m.id for m in mentions])
        )
        result = await db.execute(analysis_query)
        analyses = {a.mention_id: a for a in result.scalars().all()}
        
        # Get alerts in date range
        alert_query = select(Alert).where(
            and_(
                Alert.created_at >= start_date,
                Alert.created_at <= end_date
            )
        )
        result = await db.execute(alert_query)
        alerts = result.scalars().all()
        
        # Calculate metrics
        total_mentions = len(mentions)
        
        sentiment_counts = {
            "positive": 0,
            "neutral": 0,
            "negative_low": 0,
            "negative_medium": 0,
            "negative_high": 0
        }
        
        risk_distribution = {
            "low": 0,      # 0-30
            "medium": 0,   # 31-60
            "high": 0,     # 61-80
            "critical": 0  # 81-100
        }
        
        crisis_levels = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        
        total_risk_score = 0
        high_risk_mentions = []
        
        for mention in mentions:
            analysis = analyses.get(mention.id)
            if analysis:
                # Sentiment
                sentiment_counts[analysis.sentiment.value] += 1
                
                # Risk distribution
                if analysis.risk_score <= 30:
                    risk_distribution["low"] += 1
                elif analysis.risk_score <= 60:
                    risk_distribution["medium"] += 1
                elif analysis.risk_score <= 80:
                    risk_distribution["high"] += 1
                else:
                    risk_distribution["critical"] += 1
                
                # Crisis level
                crisis_levels[analysis.crisis_level] += 1
                
                # Total risk
                total_risk_score += analysis.risk_score
                
                # High risk mentions
                if analysis.risk_score >= 70:
                    high_risk_mentions.append({
                        "mention": mention,
                        "analysis": analysis
                    })
        
        avg_risk_score = total_risk_score / total_mentions if total_mentions > 0 else 0
        
        # Alert metrics
        alert_severity_counts = {
            "low": 0,
            "medium": 0,
            "high": 0,
            "critical": 0
        }
        
        for alert in alerts:
            alert_severity_counts[alert.severity.value] += 1
        
        # Top sources
        source_mention_counts = {}
        for mention in mentions:
            source_id = mention.source_id
            source_mention_counts[source_id] = source_mention_counts.get(source_id, 0) + 1
        
        # Get source details
        top_source_ids = sorted(source_mention_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        top_sources = []
        for source_id, count in top_source_ids:
            result = await db.execute(select(Source).where(Source.id == source_id))
            source = result.scalar_one_or_none()
            if source:
                top_sources.append({
                    "source": source,
                    "mention_count": count
                })
        
        # Sort high risk mentions by risk score
        high_risk_mentions.sort(key=lambda x: x["analysis"].risk_score, reverse=True)
        
        return {
            "report": report,
            "period": {
                "start": start_date,
                "end": end_date,
                "days": (end_date - start_date).days + 1
            },
            "summary": {
                "total_mentions": total_mentions,
                "total_alerts": len(alerts),
                "avg_risk_score": avg_risk_score,
                "high_risk_count": len(high_risk_mentions)
            },
            "sentiment_counts": sentiment_counts,
            "risk_distribution": risk_distribution,
            "crisis_levels": crisis_levels,
            "alert_severity_counts": alert_severity_counts,
            "top_sources": top_sources,
            "high_risk_mentions": high_risk_mentions[:20],  # Top 20
            "all_mentions": mentions,
            "all_alerts": alerts
        }
    
    async def _generate_pdf(
        self,
        report: Report,
        data: Dict[str, Any]
    ) -> str:
        """Generate PDF report"""
        # Create reports directory
        os.makedirs("reports", exist_ok=True)
        
        filename = f"reports/report_{report.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        doc = SimpleDocTemplate(filename, pagesize=A4)
        story = []
        
        # Title
        title = Paragraph(
            f"BÁO CÁO GIÁM SÁT DANH TIẾNG<br/>{report.title}",
            self.styles['CustomTitle']
        )
        story.append(title)
        story.append(Spacer(1, 0.3*inch))
        
        # Period
        period_text = f"Thời gian: {data['period']['start'].strftime('%d/%m/%Y')} - {data['period']['end'].strftime('%d/%m/%Y')} ({data['period']['days']} ngày)"
        story.append(Paragraph(period_text, self.styles['Normal']))
        story.append(Spacer(1, 0.2*inch))
        
        # Summary section
        story.append(Paragraph("TỔNG QUAN", self.styles['CustomHeading']))
        
        summary_data = [
            ['Chỉ số', 'Giá trị'],
            ['Tổng số mentions', str(data['summary']['total_mentions'])],
            ['Tổng số cảnh báo', str(data['summary']['total_alerts'])],
            ['Risk score trung bình', f"{data['summary']['avg_risk_score']:.1f}/100"],
            ['Mentions rủi ro cao', str(data['summary']['high_risk_count'])]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Sentiment distribution
        story.append(Paragraph("PHÂN BỐ CẢM XÚC", self.styles['CustomHeading']))
        
        sentiment_data = [
            ['Cảm xúc', 'Số lượng', 'Tỷ lệ'],
            ['Tích cực', str(data['sentiment_counts']['positive']), 
             f"{data['sentiment_counts']['positive']/data['summary']['total_mentions']*100:.1f}%" if data['summary']['total_mentions'] > 0 else "0%"],
            ['Trung lập', str(data['sentiment_counts']['neutral']),
             f"{data['sentiment_counts']['neutral']/data['summary']['total_mentions']*100:.1f}%" if data['summary']['total_mentions'] > 0 else "0%"],
            ['Tiêu cực thấp', str(data['sentiment_counts']['negative_low']),
             f"{data['sentiment_counts']['negative_low']/data['summary']['total_mentions']*100:.1f}%" if data['summary']['total_mentions'] > 0 else "0%"],
            ['Tiêu cực trung bình', str(data['sentiment_counts']['negative_medium']),
             f"{data['sentiment_counts']['negative_medium']/data['summary']['total_mentions']*100:.1f}%" if data['summary']['total_mentions'] > 0 else "0%"],
            ['Tiêu cực cao', str(data['sentiment_counts']['negative_high']),
             f"{data['sentiment_counts']['negative_high']/data['summary']['total_mentions']*100:.1f}%" if data['summary']['total_mentions'] > 0 else "0%"]
        ]
        
        sentiment_table = Table(sentiment_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
        sentiment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(sentiment_table)
        story.append(Spacer(1, 0.3*inch))
        
        # High risk mentions
        if data['high_risk_mentions']:
            story.append(PageBreak())
            story.append(Paragraph("MENTIONS RỦI RO CAO", self.styles['CustomHeading']))
            
            for item in data['high_risk_mentions'][:10]:  # Top 10
                mention = item['mention']
                analysis = item['analysis']
                
                risk_text = f"<b>Risk Score: {analysis.risk_score:.0f}/100 | Crisis Level: {analysis.crisis_level}/5</b>"
                story.append(Paragraph(risk_text, self.styles['Normal']))
                
                title_text = f"<b>{mention.title or 'No title'}</b>"
                story.append(Paragraph(title_text, self.styles['Normal']))
                
                summary_text = analysis.summary_vi or "No summary"
                story.append(Paragraph(summary_text, self.styles['Normal']))
                
                url_text = f"<i>{mention.url}</i>"
                story.append(Paragraph(url_text, self.styles['Normal']))
                
                story.append(Spacer(1, 0.2*inch))
        
        # Build PDF
        doc.build(story)
        
        return filename
    
    async def _generate_excel(
        self,
        report: Report,
        data: Dict[str, Any]
    ) -> str:
        """Generate Excel report"""
        # Create reports directory
        os.makedirs("reports", exist_ok=True)
        
        filename = f"reports/report_{report.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Tổng Quan"
        
        # Header
        ws_summary['A1'] = report.title
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f"Thời gian: {data['period']['start'].strftime('%d/%m/%Y')} - {data['period']['end'].strftime('%d/%m/%Y')}"
        
        # Summary metrics
        ws_summary['A4'] = "Chỉ Số"
        ws_summary['B4'] = "Giá Trị"
        ws_summary['A4'].font = Font(bold=True)
        ws_summary['B4'].font = Font(bold=True)
        
        ws_summary['A5'] = "Tổng số mentions"
        ws_summary['B5'] = data['summary']['total_mentions']
        
        ws_summary['A6'] = "Tổng số cảnh báo"
        ws_summary['B6'] = data['summary']['total_alerts']
        
        ws_summary['A7'] = "Risk score trung bình"
        ws_summary['B7'] = f"{data['summary']['avg_risk_score']:.1f}"
        
        ws_summary['A8'] = "Mentions rủi ro cao"
        ws_summary['B8'] = data['summary']['high_risk_count']
        
        # Sentiment distribution
        ws_summary['A10'] = "Phân Bố Cảm Xúc"
        ws_summary['A10'].font = Font(bold=True)
        
        ws_summary['A11'] = "Tích cực"
        ws_summary['B11'] = data['sentiment_counts']['positive']
        
        ws_summary['A12'] = "Trung lập"
        ws_summary['B12'] = data['sentiment_counts']['neutral']
        
        ws_summary['A13'] = "Tiêu cực thấp"
        ws_summary['B13'] = data['sentiment_counts']['negative_low']
        
        ws_summary['A14'] = "Tiêu cực trung bình"
        ws_summary['B14'] = data['sentiment_counts']['negative_medium']
        
        ws_summary['A15'] = "Tiêu cực cao"
        ws_summary['B15'] = data['sentiment_counts']['negative_high']
        
        # All mentions sheet
        ws_mentions = wb.create_sheet("Tất Cả Mentions")
        
        headers = ['ID', 'Tiêu đề', 'Nội dung', 'URL', 'Ngày đăng', 'Risk Score', 'Sentiment', 'Crisis Level']
        ws_mentions.append(headers)
        
        # Style header
        for cell in ws_mentions[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add mention data
        for mention in data['all_mentions']:
            analysis = next((a for a in data.get('analyses', []) if a.mention_id == mention.id), None)
            
            row = [
                mention.id,
                mention.title or '',
                mention.content[:200] + '...' if len(mention.content) > 200 else mention.content,
                mention.url,
                mention.published_at.strftime('%d/%m/%Y %H:%M') if mention.published_at else '',
                f"{analysis.risk_score:.1f}" if analysis else '',
                analysis.sentiment.value if analysis else '',
                analysis.crisis_level if analysis else ''
            ]
            ws_mentions.append(row)
        
        # Adjust column widths
        ws_mentions.column_dimensions['A'].width = 10
        ws_mentions.column_dimensions['B'].width = 30
        ws_mentions.column_dimensions['C'].width = 50
        ws_mentions.column_dimensions['D'].width = 40
        ws_mentions.column_dimensions['E'].width = 20
        ws_mentions.column_dimensions['F'].width = 15
        ws_mentions.column_dimensions['G'].width = 20
        ws_mentions.column_dimensions['H'].width = 15
        
        # Save workbook
        wb.save(filename)
        
        return filename


# Singleton instance
report_service = ReportService()
