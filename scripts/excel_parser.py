#!/usr/bin/env python3
"""
MHC Excel Parser
A specialized parser for the MHC pricing Excel file
"""

import openpyxl
import json
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional
import re

class MHCExcelParser:
    """Parser for MHC pricing Excel files"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.data = {}
        
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            return True
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def clean_price(self, price_str: str) -> Optional[Dict[str, Any]]:
        """Clean and parse price strings"""
        if not price_str or price_str in ['Thoả thuận', 'Thỏa thuận', 'Tặng kèm']:
            return {
                'type': 'negotiable' if 'thuận' in str(price_str) else 'free',
                'min_price': None,
                'max_price': None,
                'raw': str(price_str)
            }
        
        # Extract numbers from price string (handle Vietnamese number format with dots)
        # Pattern matches numbers with dots as thousand separators
        numbers = re.findall(r'[\d.]+(?:\.\d{3})*', str(price_str))
        if numbers:
            try:
                # Convert Vietnamese number format to integers
                prices = []
                for num in numbers:
                    # Remove dots (thousand separators) and convert to int
                    clean_num = num.replace('.', '')
                    if clean_num.isdigit():
                        prices.append(int(clean_num))
                
                if len(prices) == 1:
                    return {
                        'type': 'fixed',
                        'min_price': prices[0],
                        'max_price': prices[0],
                        'raw': str(price_str)
                    }
                elif len(prices) >= 2:
                    return {
                        'type': 'range',
                        'min_price': min(prices),
                        'max_price': max(prices),
                        'raw': str(price_str)
                    }
            except ValueError:
                pass
        
        return {
            'type': 'unknown',
            'min_price': None,
            'max_price': None,
            'raw': str(price_str)
        }
    
    def parse_service_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """Parse a service pricing sheet"""
        if not self.workbook:
            return {}
        
        sheet = self.workbook[sheet_name]
        
        # Find header row (usually around row 9)
        header_row = None
        headers = []
        
        for row_idx in range(1, 15):  # Check first 15 rows
            row_data = [sheet.cell(row_idx, col).value for col in range(1, 15)]
            text_count = sum(1 for cell in row_data if cell and isinstance(cell, str) and len(str(cell)) > 2)
            
            if text_count >= 5:  # Row with many text cells is likely header
                header_row = row_idx
                headers = [cell for cell in row_data if cell]
                break
        
        if not header_row:
            return {'error': f'Could not find header row in sheet {sheet_name}'}
        
        # Parse data rows
        services = []
        current_category = None
        
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row_data = []
            for col_idx in range(1, len(headers) + 1):
                cell_value = sheet.cell(row_idx, col_idx).value
                row_data.append(cell_value)
            
            # Skip empty rows
            if not any(cell for cell in row_data if cell):
                continue
            
            # Check if this is a category header
            first_cell = row_data[0]
            if first_cell and isinstance(first_cell, str) and not str(first_cell).isdigit():
                if len([cell for cell in row_data[1:] if cell]) == 0:  # Only first cell has data
                    current_category = first_cell
                    continue
            
            # Parse service data
            service_data = {}
            for i, header in enumerate(headers):
                if i < len(row_data):
                    value = row_data[i]
                    
                    # Special handling for price columns
                    if 'giá' in str(header).lower() and value:
                        service_data[header] = self.clean_price(value)
                    else:
                        service_data[header] = value
            
            # Add category if available
            if current_category:
                service_data['category'] = current_category
            
            # Only add if we have meaningful data
            if any(service_data.get(h) for h in headers[1:3]):  # Check if service name or details exist
                services.append(service_data)
        
        return {
            'sheet_name': sheet_name,
            'header_row': header_row,
            'headers': headers,
            'services': services,
            'total_services': len(services)
        }
    
    def parse_all_sheets(self) -> Dict[str, Any]:
        """Parse all sheets in the workbook"""
        if not self.load_workbook():
            return {'error': 'Failed to load workbook'}
        
        result = {
            'file_info': {
                'filename': Path(self.file_path).name,
                'sheet_names': self.workbook.sheetnames,
                'total_sheets': len(self.workbook.sheetnames)
            },
            'sheets': {}
        }
        
        for sheet_name in self.workbook.sheetnames:
            print(f"Parsing sheet: {sheet_name}")
            result['sheets'][sheet_name] = self.parse_service_sheet(sheet_name)
        
        self.workbook.close()
        return result
    
    def export_to_json(self, output_path: str, data: Dict[str, Any]):
        """Export parsed data to JSON"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    
    def export_to_csv(self, output_dir: str, data: Dict[str, Any]):
        """Export each sheet to separate CSV files"""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        for sheet_name, sheet_data in data['sheets'].items():
            if 'services' in sheet_data:
                # Flatten the data for CSV
                csv_data = []
                for service in sheet_data['services']:
                    flat_service = {}
                    for key, value in service.items():
                        if isinstance(value, dict) and 'raw' in value:
                            # For price data, use the raw value
                            flat_service[key] = value['raw']
                            flat_service[f"{key}_min"] = value.get('min_price')
                            flat_service[f"{key}_max"] = value.get('max_price')
                            flat_service[f"{key}_type"] = value.get('type')
                        else:
                            flat_service[key] = value
                    csv_data.append(flat_service)
                
                if csv_data:
                    df = pd.DataFrame(csv_data)
                    csv_filename = f"{sheet_name.replace('/', '_').replace(' ', '_')}.csv"
                    csv_path = output_path / csv_filename
                    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                    print(f"Exported {sheet_name} to {csv_path}")

def main():
    """Main function to demonstrate usage"""
    parser = MHCExcelParser("data/mhc_bao_gia_khung_xlkh.xlsx")
    
    print("Parsing MHC Excel file...")
    data = parser.parse_all_sheets()
    
    if 'error' in data:
        print(f"Error: {data['error']}")
        return
    
    # Export to JSON
    parser.export_to_json("data/mhc_parsed_detailed.json", data)
    print("Exported detailed data to: data/mhc_parsed_detailed.json")
    
    # Export to CSV
    parser.export_to_csv("data/csv_exports", data)
    print("Exported CSV files to: data/csv_exports/")
    
    # Print summary
    print(f"\nSummary:")
    print(f"Total sheets: {data['file_info']['total_sheets']}")
    for sheet_name, sheet_data in data['sheets'].items():
        if 'total_services' in sheet_data:
            print(f"  - {sheet_name}: {sheet_data['total_services']} services")

if __name__ == "__main__":
    main()