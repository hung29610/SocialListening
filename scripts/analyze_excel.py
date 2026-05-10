#!/usr/bin/env python3
"""
Excel File Analyzer
Analyzes the MHC Excel file structure and content
"""

import openpyxl
import json
from pathlib import Path
from typing import Dict, List, Any
import re

def clean_cell_value(value):
    """Clean and normalize cell values"""
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value

def detect_header_row(sheet, max_rows_to_check=10):
    """Detect which row contains headers by looking for text-heavy rows"""
    header_candidates = []
    
    for row_idx in range(1, min(max_rows_to_check + 1, sheet.max_row + 1)):
        row_data = []
        text_count = 0
        
        for col_idx in range(1, min(20, sheet.max_column + 1)):  # Check first 20 columns
            cell_value = clean_cell_value(sheet.cell(row_idx, col_idx).value)
            row_data.append(cell_value)
            
            if cell_value and isinstance(cell_value, str) and len(cell_value) > 2:
                text_count += 1
        
        if text_count >= 3:  # Row with at least 3 text cells likely to be header
            header_candidates.append((row_idx, row_data, text_count))
    
    if header_candidates:
        # Return the row with most text cells
        return max(header_candidates, key=lambda x: x[2])
    
    return None

def analyze_sheet(sheet):
    """Analyze a single sheet"""
    analysis = {
        'name': sheet.title,
        'dimensions': {
            'max_row': sheet.max_row,
            'max_column': sheet.max_column
        },
        'headers': [],
        'header_row': None,
        'sample_data': [],
        'column_analysis': {},
        'summary': ''
    }
    
    # Detect header row
    header_info = detect_header_row(sheet)
    if header_info:
        header_row_idx, headers, _ = header_info
        analysis['header_row'] = header_row_idx
        analysis['headers'] = [h for h in headers if h is not None]
    
    # Get sample data (first 5 rows after header)
    start_row = (analysis['header_row'] or 0) + 1
    for row_idx in range(start_row, min(start_row + 5, sheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(len(analysis['headers']) + 1, sheet.max_column + 1)):
            cell_value = clean_cell_value(sheet.cell(row_idx, col_idx).value)
            row_data.append(cell_value)
        
        if any(cell for cell in row_data if cell is not None):
            analysis['sample_data'].append(row_data)
    
    # Analyze columns
    if analysis['headers']:
        for col_idx, header in enumerate(analysis['headers'], 1):
            if header:
                col_values = []
                for row_idx in range(start_row, min(start_row + 20, sheet.max_row + 1)):
                    value = clean_cell_value(sheet.cell(row_idx, col_idx).value)
                    if value is not None:
                        col_values.append(value)
                
                analysis['column_analysis'][header] = {
                    'sample_values': col_values[:5],
                    'total_values': len(col_values),
                    'data_types': list(set(type(v).__name__ for v in col_values))
                }
    
    # Generate summary
    if analysis['headers']:
        analysis['summary'] = f"Sheet '{sheet.title}' contains {len(analysis['headers'])} columns with headers: {', '.join(str(h) for h in analysis['headers'][:5])}{'...' if len(analysis['headers']) > 5 else ''}. Data starts at row {start_row} with {sheet.max_row - start_row + 1} data rows."
    else:
        analysis['summary'] = f"Sheet '{sheet.title}' has {sheet.max_row} rows and {sheet.max_column} columns. No clear headers detected."
    
    return analysis

def analyze_excel_file(file_path: str) -> Dict[str, Any]:
    """Analyze the entire Excel file"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        analysis = {
            'file_info': {
                'filename': Path(file_path).name,
                'sheet_count': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames
            },
            'sheets': {},
            'summary': ''
        }
        
        # Analyze each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            analysis['sheets'][sheet_name] = analyze_sheet(sheet)
        
        # Generate overall summary
        total_rows = sum(sheet_data['dimensions']['max_row'] for sheet_data in analysis['sheets'].values())
        analysis['summary'] = f"Excel file contains {len(workbook.sheetnames)} sheets with a total of {total_rows} rows across all sheets."
        
        workbook.close()
        return analysis
        
    except Exception as e:
        return {'error': f"Failed to analyze Excel file: {str(e)}"}

def create_structured_json(analysis: Dict[str, Any]) -> Dict[str, Any]:
    """Convert analysis to structured JSON format"""
    if 'error' in analysis:
        return analysis
    
    structured_data = {
        'metadata': {
            'filename': analysis['file_info']['filename'],
            'total_sheets': analysis['file_info']['sheet_count'],
            'sheet_names': analysis['file_info']['sheet_names'],
            'analysis_summary': analysis['summary']
        },
        'sheets': {}
    }
    
    for sheet_name, sheet_data in analysis['sheets'].items():
        structured_sheet = {
            'metadata': {
                'name': sheet_name,
                'dimensions': sheet_data['dimensions'],
                'header_row': sheet_data['header_row'],
                'summary': sheet_data['summary']
            },
            'structure': {
                'headers': sheet_data['headers'],
                'column_count': len(sheet_data['headers']),
                'estimated_data_rows': sheet_data['dimensions']['max_row'] - (sheet_data['header_row'] or 0)
            },
            'sample_data': {
                'headers': sheet_data['headers'],
                'rows': sheet_data['sample_data']
            },
            'column_analysis': sheet_data['column_analysis']
        }
        
        structured_data['sheets'][sheet_name] = structured_sheet
    
    return structured_data

def create_markdown_preview(analysis: Dict[str, Any]) -> str:
    """Create a markdown preview of the Excel analysis"""
    if 'error' in analysis:
        return f"# Excel Analysis Error\n\n{analysis['error']}"
    
    md_content = []
    md_content.append("# Excel File Analysis Report")
    md_content.append(f"\n**File:** {analysis['file_info']['filename']}")
    md_content.append(f"**Total Sheets:** {analysis['file_info']['sheet_count']}")
    md_content.append(f"\n## Summary\n{analysis['summary']}")
    
    md_content.append("\n## Sheet Overview")
    for i, sheet_name in enumerate(analysis['file_info']['sheet_names'], 1):
        md_content.append(f"{i}. **{sheet_name}**")
    
    md_content.append("\n## Detailed Sheet Analysis")
    
    for sheet_name, sheet_data in analysis['sheets'].items():
        md_content.append(f"\n### Sheet: {sheet_name}")
        md_content.append(f"\n{sheet_data['summary']}")
        
        md_content.append(f"\n**Dimensions:** {sheet_data['dimensions']['max_row']} rows × {sheet_data['dimensions']['max_column']} columns")
        
        if sheet_data['headers']:
            md_content.append(f"\n**Headers (Row {sheet_data['header_row']}):**")
            for i, header in enumerate(sheet_data['headers'], 1):
                md_content.append(f"{i}. {header}")
        
        if sheet_data['sample_data']:
            md_content.append(f"\n**Sample Data:**")
            md_content.append("| " + " | ".join(str(h) for h in sheet_data['headers']) + " |")
            md_content.append("| " + " | ".join("---" for _ in sheet_data['headers']) + " |")
            
            for row in sheet_data['sample_data'][:3]:  # Show first 3 rows
                row_str = "| " + " | ".join(str(cell) if cell is not None else "" for cell in row) + " |"
                md_content.append(row_str)
        
        if sheet_data['column_analysis']:
            md_content.append(f"\n**Column Analysis:**")
            for col_name, col_info in sheet_data['column_analysis'].items():
                if col_info['sample_values']:
                    sample_str = ", ".join(str(v) for v in col_info['sample_values'][:3])
                    md_content.append(f"- **{col_name}**: {col_info['total_values']} values, types: {', '.join(col_info['data_types'])}, samples: {sample_str}")
    
    md_content.append(f"\n---\n*Analysis generated on {Path(__file__).name}*")
    
    return "\n".join(md_content)

def main():
    """Main function to analyze the Excel file"""
    file_path = "data/mhc_bao_gia_khung_xlkh.xlsx"
    
    print("Analyzing Excel file...")
    analysis = analyze_excel_file(file_path)
    
    if 'error' in analysis:
        print(f"Error: {analysis['error']}")
        return
    
    print(f"Found {analysis['file_info']['sheet_count']} sheets:")
    for sheet_name in analysis['file_info']['sheet_names']:
        print(f"  - {sheet_name}")
    
    # Create structured JSON
    structured_data = create_structured_json(analysis)
    
    # Save JSON output
    json_output_path = "data/parsed_excel_summary.json"
    with open(json_output_path, 'w', encoding='utf-8') as f:
        json.dump(structured_data, f, indent=2, ensure_ascii=False)
    print(f"Saved structured data to: {json_output_path}")
    
    # Create markdown preview
    markdown_content = create_markdown_preview(analysis)
    
    # Save markdown output
    md_output_path = "data/parsed_excel_preview.md"
    with open(md_output_path, 'w', encoding='utf-8') as f:
        f.write(markdown_content)
    print(f"Saved markdown preview to: {md_output_path}")
    
    print("\nAnalysis complete!")

if __name__ == "__main__":
    main()